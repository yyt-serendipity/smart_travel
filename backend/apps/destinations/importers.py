from __future__ import annotations

from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from apps.core.tagging import normalize_public_tags
from apps.destinations.models import Attraction, TravelCity
from apps.destinations.services import (
    clean_text,
    compact_text,
    compute_city_profile,
    fit_text,
    infer_destination_type,
    infer_province,
    infer_tags,
    parse_ticket_data,
    safe_decimal,
)


EXPECTED_HEADERS = (
    "名字",
    "链接",
    "地址",
    "介绍",
    "开放时间",
    "图片链接",
    "评分",
    "建议游玩时间",
    "建议季节",
    "门票",
    "小贴士",
    "Page",
)


def _coerce_excel_source(excel_source: Path | str | bytes | BinaryIO):
    if isinstance(excel_source, (str, Path)):
        return excel_source
    if isinstance(excel_source, bytes):
        return BytesIO(excel_source)
    return excel_source


def _resolve_source_name(excel_source: Path | str | bytes | BinaryIO, source_name: str | None) -> str:
    if source_name:
        return source_name
    if isinstance(excel_source, (str, Path)):
        return Path(excel_source).name
    return "uploaded.xlsx"


def load_sheet_rows(excel_source: Path | str | bytes | BinaryIO) -> list[tuple]:
    workbook = load_workbook(_coerce_excel_source(excel_source), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    return list(worksheet.iter_rows(values_only=True))


def validate_headers(headers) -> bool:
    normalized = tuple(clean_text(item) for item in headers)
    return normalized[: len(EXPECTED_HEADERS)] == EXPECTED_HEADERS


def build_city_defaults(city_name: str, rows: list[tuple], source_file: str) -> dict:
    descriptions = [clean_text(row[3]) for row in rows if len(row) > 3 and clean_text(row[3])]
    addresses = [clean_text(row[2]) for row in rows if len(row) > 2 and clean_text(row[2])]
    seasons = [clean_text(row[8]) for row in rows if len(row) > 8 and clean_text(row[8])]
    images = [clean_text(row[5]) for row in rows if len(row) > 5 and clean_text(row[5])]
    ticket_infos = [parse_ticket_data(row[9]) for row in rows if len(row) > 9 and clean_text(row[9])]

    tag_pool = []
    for row in rows[:30]:
        tag_pool.extend(
            infer_tags(
                row[0] if len(row) > 0 else "",
                row[2] if len(row) > 2 else "",
                row[3] if len(row) > 3 else "",
            )
        )

    # Build a usable city card straight from workbook content so the city list works after one import.
    return {
        "province": fit_text(infer_province(city_name, *addresses, *descriptions), 100),
        "destination_type": infer_destination_type(city_name),
        "short_intro": fit_text(
            f"{city_name}收录了 {len(rows)} 个景点，适合城市漫游、周末度假和 AI 行程规划。",
            220,
        ),
        "overview": compact_text(
            descriptions[0] if descriptions else f"{city_name}拥有丰富的旅游资源，适合慢游体验与景点打卡。",
            800,
        ),
        "travel_highlights": "、".join(tag for tag, _ in Counter(tag_pool).most_common(4)),
        "cover_image": fit_text(images[0] if images else "", 200),
        "best_season": fit_text(Counter(seasons).most_common(1)[0][0] if seasons else "", 200),
        "average_ticket": fit_text(ticket_infos[0] if ticket_infos else "", 255),
        "tags": normalize_public_tags([tag for tag, _ in Counter(tag_pool).most_common(5)]),
        "travel_tips": "",
        "source_file": fit_text(source_file, 255),
        "is_featured": len(rows) >= 15,
    }


def import_excel_file(
    excel_source: Path | str | bytes | BinaryIO,
    source_name: str | None = None,
    overwrite: bool = True,
) -> TravelCity:
    rows = load_sheet_rows(excel_source)
    resolved_source_name = _resolve_source_name(excel_source, source_name)
    if not rows:
        raise ValueError(f"{resolved_source_name} is empty.")

    headers = rows[0]
    data_rows = [row for row in rows[1:] if row and clean_text(row[0])]
    if not validate_headers(headers):
        raise ValueError(f"{resolved_source_name} header mismatch: {headers}")

    city_name = Path(resolved_source_name).stem
    city_defaults = build_city_defaults(city_name, data_rows, resolved_source_name)
    city, _ = TravelCity.objects.update_or_create(name=city_name, defaults=city_defaults)

    seen_names = set()
    # Use (city, attraction name) as the natural key so repeated imports refresh rows instead of duplicating them.
    for row in data_rows:
        attraction_name = fit_text(row[0], 200)
        seen_names.add(attraction_name)
        defaults = {
            "source_url": fit_text(row[1], 200),
            "address": clean_text(row[2]),
            "description": clean_text(row[3]),
            "opening_hours": clean_text(row[4]),
            "image_url": fit_text(row[5], 200),
            "rating": safe_decimal(row[6]),
            "suggested_play_time": fit_text(row[7], 120),
            "best_season": fit_text(row[8], 255),
            "ticket_info": parse_ticket_data(row[9]),
            "tips": clean_text(row[10]),
            "source_page": int(row[11] or 1),
            "tags": normalize_public_tags(infer_tags(row[0], row[2], row[3])),
            "source_file": fit_text(resolved_source_name, 255),
            "imported_from_excel": True,
        }
        Attraction.objects.update_or_create(city=city, name=attraction_name, defaults=defaults)

    if overwrite:
        # In overwrite mode, the workbook becomes the source of truth for that city's attractions.
        city.attractions.exclude(name__in=seen_names).delete()

    compute_city_profile(city)
    city.save()
    return city


def import_excel_directory(directory: Path, overwrite: bool = False, limit: int | None = None) -> list[TravelCity]:
    imported = []
    files = sorted(directory.glob("*.xlsx"))
    if limit is not None:
        files = files[:limit]
    for excel_path in files:
        imported.append(import_excel_file(excel_path, overwrite=overwrite))
    return imported
